import os
import sys
import json
import datetime
import math
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import tempfile


# SQLAlchemy
from sqlalchemy import (create_engine, Column, Integer, String, Date, DateTime,
                        Float, Boolean, ForeignKey, func, Index, and_, or_)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker
from sqlalchemy.exc import SQLAlchemyError

# CLI / UI
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt, Confirm
from rich.panel import Panel
from rich import box

# Exports & charts
from openpyxl import Workbook
from fpdf import FPDF
import matplotlib.pyplot as plt

# Optional encryption
from cryptography.fernet import Fernet, InvalidToken

# ---------- Configuration ----------
DB_FILE = "expenses.db"  # chosen option 3-A
PDF_FONT_FILE = "DejaVuSans.ttf"  # place DejaVuSans.ttf next to script to support unicode in PDF
ENCRYPT_NOTES = bool(os.getenv("EXPENSE_ENCRYPT_NOTES", "0") == "1")
FERNET_KEY = os.getenv("EXPENSE_KEY")  # must be a 32 url-safe base64 key if encryption enabled

# ---------- Setup Console ----------
console = Console()

# ---------- Database Setup ----------
Base = declarative_base()
engine = create_engine(f"sqlite:///{DB_FILE}", echo=False, future=True)
SessionLocal = sessionmaker(bind=engine)

# ---------- Encryption helper ----------
fernet = None
if ENCRYPT_NOTES:
    if not FERNET_KEY:
        console.print("[red]EXPENSE_ENCRYPTION enabled but EXPENSE_KEY not set. Disable or set key.[/red]")
        ENCRYPT_NOTES = False
    else:
        try:
            if ENCRYPT_NOTES:
                if not FERNET_KEY or len(FERNET_KEY.encode()) != 44:
                    console.print("[red]Invalid EXPENSE_KEY, disabling encryption[/red]")
                    ENCRYPT_NOTES = False
                else:
                    fernet = Fernet(FERNET_KEY.encode())

        except Exception:
            console.print("[red]Invalid EXPENSE_KEY. Disable encryption or supply a valid key.[/red]")
            ENCRYPT_NOTES = False

def encrypt_note(plain: str) -> str:
    if not ENCRYPT_NOTES or not plain:
        return plain or ""
    return fernet.encrypt(plain.encode()).decode()

def decrypt_note(token: str) -> str:
    if not ENCRYPT_NOTES or not token:
        return token or ""
    try:
        return fernet.decrypt(token.encode()).decode()
    except InvalidToken:
        return "[decryption failed]"
def valid_date_str(date_str: str) -> bool:
    """Validate date format YYYY-MM-DD"""
    try:
        datetime.datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


# ---------- Models ----------
class Category(Base):
    __tablename__ = "categories"
    id = Column(Integer, primary_key=True)
    name = Column(String(100), unique=True, index=True, nullable=False)
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    budgets = relationship("Budget", back_populates="category")
    expenses = relationship("Expense", back_populates="category_obj")

class Budget(Base):
    __tablename__ = "budgets"
    id = Column(Integer, primary_key=True)
    category_id = Column(Integer, ForeignKey("categories.id", ondelete="CASCADE"), index=True)
    amount = Column(Float, nullable=False)  # budget amount per month
    currency = Column(String(8), default="INR")
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    category = relationship("Category", back_populates="budgets")

class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True)
    amount = Column(Float, nullable=False, index=True)
    currency = Column(String(8), default="INR")
    note = Column(String, nullable=True)  # can be encrypted
    date = Column(Date, index=True, default=datetime.date.today)
    category_id = Column(Integer, ForeignKey("categories.id", ondelete="SET NULL"), index=True, nullable=True)
    created_at = Column(DateTime, default=datetime.datetime.utcnow)
    deleted = Column(Boolean, default=False, index=True)  # soft delete flag

    category_obj = relationship("Category", back_populates="expenses")

    # index composite for frequent queries
Index('ix_expense_date_amount', Expense.date, Expense.amount)

class ExpenseHistory(Base):
    __tablename__ = "expense_history"
    id = Column(Integer, primary_key=True)
    expense_id = Column(Integer, index=True)
    action = Column(String(30))  # "create","update","delete"
    snapshot = Column(String)  # JSON snapshot
    timestamp = Column(DateTime, default=datetime.datetime.utcnow)

# A simple schema version row for manual migrations
class MetaInfo(Base):
    __tablename__ = "meta_info"
    key = Column(String(50), primary_key=True)
    value = Column(String(200))

# Create tables if not present
def init_db():
    Base.metadata.create_all(bind=engine)
    # set schema version if missing
    session = SessionLocal()
    try:
        v = session.get(MetaInfo, "schema_version")
        if not v:
            session.add(MetaInfo(key="schema_version", value="1"))
            session.commit()
    except Exception:
        session.rollback()
    finally:
        session.close()

init_db()

# ---------- Utility DB functions ----------
def create_category_if_missing(session, name: str):
    name = name.strip().title()
    cat = session.query(Category).filter(func.lower(Category.name) == name.lower()).first()
    if not cat:
        cat = Category(name=name)
        session.add(cat)
        session.commit()
    return cat

def add_history(session, expense_obj, action: str):
    try:
        snapshot = {
            "id": expense_obj.id,
            "amount": expense_obj.amount,
            "currency": expense_obj.currency,
            "note": expense_obj.note,
            "date": expense_obj.date.isoformat() if expense_obj.date else None,
            "category_id": expense_obj.category_id,
            "deleted": bool(expense_obj.deleted),
            "created_at": expense_obj.created_at.isoformat() if expense_obj.created_at else None
        }
        hist = ExpenseHistory(expense_id=expense_obj.id, action=action, snapshot=json.dumps(snapshot))
        session.add(hist)
        session.commit()
    except Exception:
        session.rollback()

# ---------- Core operations ----------
def add_expense_db(amount: float, note: str, date_str: str, category_name: str, currency: str = "INR"):
    session = SessionLocal()
    try:
        # validate
        try:
            amount_v = float(amount)
            if amount_v <= 0:
                console.print("[red]Amount must be greater than 0[/red]")
                return
        except (ValueError, TypeError):
            console.print("[red]Invalid amount value[/red]")
            return

        # date
        if date_str and valid_date_str(date_str):
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        else:
            date_obj = datetime.date.today()
            if date_str:
                console.print("[yellow]Invalid date format, using today[/yellow]")


        # category
        cat = None
        if category_name:
            cat = create_category_if_missing(session, category_name)
        # encrypt note if enabled
        note_stored = encrypt_note(note) if ENCRYPT_NOTES else (note or "")

        expense = Expense(amount=round(amount_v, 2),
                          currency=currency or "INR",
                          note=note_stored,
                          date=date_obj,
                          category_id=cat.id if cat else None)
        session.add(expense)
        session.commit()

        add_history(session, expense, "create")

        # budget alert (check monthly total for that category)
        if cat:
            budget = session.query(Budget).filter(Budget.category_id == cat.id).order_by(Budget.created_at.desc()).first()
            if budget:
                # calculate month total including this expense
                month_start = date_obj.replace(day=1)
                import calendar  # add at top with imports if not present
                last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
                month_end = date_obj.replace(day=last_day)

                total = session.query(func.sum(Expense.amount)).filter(
                    Expense.category_id == cat.id,
                    Expense.deleted == False,
                    Expense.date >= month_start,
                    Expense.date <= month_end
                ).scalar() or 0.0
                if total >= budget.amount:
                    console.print(f"[red]⚠ Budget exceeded for category {cat.name} (budget ₹{budget.amount:.2f})[/red]")
                elif total >= 0.8 * budget.amount:
                    console.print(f"[yellow]⚠ Approaching budget for {cat.name}: {total:.2f}/ {budget.amount:.2f}[/yellow]")

        console.print("[green]✔ Expense saved (id: %s)[/green]" % expense.id)
    except SQLAlchemyError as e:
        session.rollback()
        console.print(f"[red]DB Error: {e}[/red]")
    finally:
        session.close()

def view_expenses_db(page:int=1, per_page:int=20, sort_by:str="date", desc:bool=True, include_deleted:bool=False):
    session = SessionLocal()
    try:
        q = session.query(Expense)
        if not include_deleted:
            q = q.filter(Expense.deleted == False)
        if sort_by == "date":
            q = q.order_by(Expense.date.desc() if desc else Expense.date)
        elif sort_by == "amount":
            q = q.order_by(Expense.amount.desc() if desc else Expense.amount)
        elif sort_by == "id":
            q = q.order_by(Expense.id.desc() if desc else Expense.id)

        total = q.count()
        pages = math.ceil(total / per_page) if per_page else 1
        items = q.offset((page-1)*per_page).limit(per_page).all()

        # prepare rows
        rows = []
        for e in items:
            cat = None
            if e.category_id:
                cat_obj = session.get(Category, e.category_id)
                cat = cat_obj.name if cat_obj else ""
            note = decrypt_note(e.note) if ENCRYPT_NOTES else (e.note or "")
            rows.append({
                "id": e.id, "amount": e.amount, "currency": e.currency,
                "category": cat, "date": e.date.isoformat() if e.date else "", "note": note, "deleted": e.deleted
            })
        return {"total": total, "pages": pages, "page": page, "per_page": per_page, "items": rows}
    except SQLAlchemyError as exc:
        console.print(f"[red]DB error: {exc}[/red]")
        return {"total":0,"pages":0,"page":1,"per_page":per_page,"items":[]}
    finally:
        session.close()

def delete_expense_db(eid:int, soft=True):
    session = SessionLocal()
    try:
        e = session.get(Expense, eid)
        if not e:
            console.print("[red]ID not found[/red]")
            return
        if soft:
            e.deleted = True
            session.commit()
            add_history(session, e, "delete")
            console.print(f"[green]Soft-deleted ID {eid}[/green]")
        else:
            session.delete(e)
            session.commit()
            console.print(f"[green]Hard-deleted ID {eid}[/green]")
    except SQLAlchemyError as exc:
        session.rollback()
        console.print(f"[red]DB error: {exc}[/red]")
    finally:
        session.close()

def update_expense_db(eid:int, amount=None, note=None, date_str=None, category_name=None, currency=None):
    session = SessionLocal()
    try:
        e = session.get(Expense, eid)
        if not e:
            console.print("[red]ID not found[/red]")
            return
        old_snapshot = json.dumps({
            "amount": e.amount, "note": e.note, "date": e.date.isoformat() if e.date else None,
            "category_id": e.category_id, "currency": e.currency
        })
        changed = False
        if amount is not None:
            e.amount = float(amount)
            changed = True
        if note is not None:
            e.note = encrypt_note(note) if ENCRYPT_NOTES else note
            changed = True
        if date_str:
            try:
                e.date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                changed = True
            except Exception:
                console.print("[yellow]Invalid date ignored[/yellow]")
        if category_name:
            cat = create_category_if_missing(session, category_name)
            e.category_id = cat.id
            changed = True
        if currency:
            e.currency = currency
            changed = True
        if changed:
            session.commit()
            add_history(session, e, "update")
            console.print("[green]Updated[/green]")
        else:
            console.print("[yellow]No changes made[/yellow]")
    except SQLAlchemyError as exc:
        session.rollback()
        console.print(f"[red]DB error: {exc}[/red]")
    finally:
        session.close()

# ---------- Search / Filters ----------
def search_expenses(keyword: str = None, amount_min: float=None, amount_max:float=None,
                    date_from: str=None, date_to: str=None, category: str=None,
                    page:int=1, per_page:int=20):
    session = SessionLocal()
    try:
        q = session.query(Expense).filter(Expense.deleted==False)
        if keyword:
            term = f"%{keyword}%"
            # search in note (encrypted notes can't be searched reliably if encrypted)
            if ENCRYPT_NOTES:
                console.print("[yellow]Note searching disabled when notes are encrypted.[/yellow]")
            else:
                q = q.filter(Expense.note.ilike(term))
        if amount_min is not None:
            q = q.filter(Expense.amount >= amount_min)
        if amount_max is not None:
            q = q.filter(Expense.amount <= amount_max)
        if date_from:
            try:
                dfrom = datetime.datetime.strptime(date_from, "%Y-%m-%d").date()
                q = q.filter(Expense.date >= dfrom)
            except Exception:
                pass
        if date_to:
            try:
                dto = datetime.datetime.strptime(date_to, "%Y-%m-%d").date()
                q = q.filter(Expense.date <= dto)
            except Exception:
                pass
        if category:
            # join category
            cat = session.query(Category).filter(func.lower(Category.name)==category.lower()).first()
            if cat:
                q = q.filter(Expense.category_id == cat.id)
            else:
                return {"total":0,"pages":0,"page":1,"per_page":per_page,"items":[]}

        total = q.count()
        pages = math.ceil(total / per_page) if per_page else 1
        items = q.offset((page-1)*per_page).limit(per_page).all()
        rows = []
        for e in items:
            cat_name = ""
            if e.category_id:
                c = session.get(Category, e.category_id)
                cat_name = c.name if c else ""
            note_plain = decrypt_note(e.note) if ENCRYPT_NOTES else (e.note or "")
            rows.append({
                "id": e.id, "amount": e.amount, "currency": e.currency,
                "category": cat_name, "date": e.date.isoformat() if e.date else "", "note": note_plain
            })
        return {"total":total,"pages":pages,"page":page,"per_page":per_page,"items":rows}

    except SQLAlchemyError as exc:
        console.print(f"[red]DB error: {exc}[/red]")
        return {"total":0,"pages":0,"page":1,"per_page":per_page,"items":[]}
    finally:
        session.close()

# ---------- Analytics ----------
def spending_trend(period_days: int = 30):
    """Return daily totals for last period_days days (list of (date,total))."""
    session = SessionLocal()
    try:
        today = datetime.date.today()
        start = today - datetime.timedelta(days=period_days-1)
        rows = session.query(Expense.date, func.sum(Expense.amount)).filter(
            Expense.deleted==False,
            Expense.date >= start,
            Expense.date <= today
        ).group_by(Expense.date).order_by(Expense.date).all()
        # produce continuous list
        totals_map = {r[0].isoformat(): float(r[1]) for r in rows}
        result = []
        for i in range(period_days):
            d = (start + datetime.timedelta(days=i)).isoformat()
            result.append((d, totals_map.get(d, 0.0)))
        return result
    finally:
        session.close()

# ---------- Exports (Excel + PDF with charts) ----------
def export_to_excel_rows(rows, filename=None):
    if not rows:
        console.print("[yellow]No data to export.[/yellow]")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    headers = ["ID", "Amount", "Currency", "Category", "Date", "Note"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get("id"), float(r.get("amount",0)), r.get("currency",""), r.get("category",""), r.get("date",""), r.get("note","")])
    if not filename:
        filename = f"expenses_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        wb.save(filename)
        console.print(f"[green]✔ Excel exported:[/green] [cyan]{filename}[/cyan]")
    except Exception as ex:
        console.print(f"[red]Failed to save Excel: {ex}[/red]")

from fpdf.enums import XPos, YPos  # add this at the top of your imports

def export_to_pdf_rows(rows, filename=None, embed_chart=False, chart_days=30):
    if not rows and not embed_chart:
        console.print("[yellow]No data to export.[/yellow]")
        return
    if not filename:
        filename = f"expenses_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ---------- Title ----------
    pdf.set_font("Helvetica", "B", 14)
    title = f"Expense Report - {datetime.date.today().isoformat()}"
    pdf.cell(0, 10, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(5)

    if rows:
        # ---------- Table Header ----------
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(15, 8, "ID", border=1)
        pdf.cell(30, 8, "Amount", border=1)
        pdf.cell(25, 8, "Curr", border=1)
        pdf.cell(40, 8, "Category", border=1)
        pdf.cell(25, 8, "Date", border=1)
        pdf.cell(55, 8, "Note", border=1)
        pdf.ln()

        # ---------- Table Rows ----------
        pdf.set_font("Helvetica", size=9)
        for r in rows:
            note = r.get("note","") or ""
            pdf.cell(15, 8, str(r.get("id","")), border=1)
            pdf.cell(30, 8, f"Rs.{float(r.get('amount',0)):.2f}", border=1)
            pdf.cell(25, 8, str(r.get("currency","")), border=1)
            pdf.cell(40, 8, str(r.get("category",""))[:20], border=1)
            pdf.cell(25, 8, str(r.get("date","")), border=1)
            pdf.cell(55, 8, note[:40], border=1)
            pdf.ln()

    # ---------- Optional Trend Chart ----------
    if embed_chart:
        trend = spending_trend(period_days=chart_days)
        dates = [t[0] for t in trend]
        totals = [t[1] for t in trend]
        tmpfile = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        plt.figure(figsize=(8,2.5))
        plt.plot(dates, totals, marker='o', linewidth=1)
        plt.xticks(rotation=45, fontsize=6)
        plt.tight_layout()
        plt.savefig(tmpfile.name, dpi=150)
        plt.close()

        pdf.add_page()
        pdf.image(tmpfile.name, x=15, w=180)

        try:
            os.unlink(tmpfile.name)
        except:
            pass

    try:
        pdf.output(filename)
        console.print(f"[green]✔ PDF exported:[/green] [cyan]{filename}[/cyan]")
    except Exception as ex:
        console.print(f"[red]Failed to save PDF: {ex}[/red]")


# ---------- CLI Helpers ----------
def prompt_add():
    amount = Prompt.ask("Amount (₹)")
    try:
        amt = float(amount)
    except:
        console.print("[red]Invalid amount[/red]")
        return
    note = Prompt.ask("Note", default="")
    date = Prompt.ask("Date (YYYY-MM-DD) or Enter", default="")
    date_valid = date if valid_date_str(date) else datetime.date.today().isoformat()
    if date != date_valid:
        console.print("[yellow]Invalid date, using today[/yellow]")
    category = Prompt.ask("Category", default="Other")
    currency = Prompt.ask("Currency", default="INR")
    add_expense_db(amt, note, date_valid, category, currency)

def prompt_view():
    page = int(Prompt.ask("Page", default="1"))
    per_page = int(Prompt.ask("Per page", default="20"))
    sort = Prompt.ask("Sort by", choices=["date","amount","id"], default="date")
    desc = Confirm.ask("Descending?", default=True)
    res = view_expenses_db(page=page, per_page=per_page, sort_by=sort, desc=desc)
    table = Table(box=box.SIMPLE, show_header=True, header_style="bold magenta")
    table.add_column("No.", style="dim", width=5, justify="right")
    table.add_column("ID", style="cyan", width=6)
    table.add_column("Amount (₹)", justify="right")
    table.add_column("Curr")
    table.add_column("Category")
    table.add_column("Date")
    table.add_column("Note")
    for idx, it in enumerate(res["items"], start=1 + (res["page"]-1)*res["per_page"]):
        table.add_row(str(idx), str(it["id"]), f"{float(it['amount']):.2f}", it.get("currency",""), it.get("category",""), it.get("date",""), str(it.get("note",""))[:40])
    console.print(table)
    console.print(f"Page {res['page']}/{res['pages']} — total {res['total']}")

def prompt_delete():
    eid = int(Prompt.ask("Expense ID to soft-delete"))
    delete_expense_db(eid, soft=True)

def prompt_update():
    eid = int(Prompt.ask("Expense ID to update"))
    # ask optionally fields
    amt = Prompt.ask("New amount (leave blank to skip)", default="")
    amt_v = float(amt) if amt.strip() else None
    note = Prompt.ask("New note (leave blank to skip)", default="")
    note_v = note if note.strip() else None
    date = Prompt.ask("New date (YYYY-MM-DD) or blank", default="")
    date_v = date if date.strip() else None
    cat = Prompt.ask("New category (blank skip)", default="")
    cat_v = cat if cat.strip() else None
    curr = Prompt.ask("Currency (blank skip)", default="")
    curr_v = curr if curr.strip() else None
    update_expense_db(eid, amount=amt_v, note=note_v, date_str=date_v, category_name=cat_v, currency=curr_v)

def prompt_search_export():
    kw = Prompt.ask("Keyword (search note) or blank", default="")
    amt_min = Prompt.ask("Min amount (blank skip)", default="")
    amt_max = Prompt.ask("Max amount (blank skip)", default="")
    df = Prompt.ask("Date from YYYY-MM-DD (blank skip)", default="")
    dt = Prompt.ask("Date to YYYY-MM-DD (blank skip)", default="")
    cat = Prompt.ask("Category (blank skip)", default="")
    page = int(Prompt.ask("Page", default="1"))
    per = int(Prompt.ask("Per page", default="1000"))
    res = search_expenses(keyword=kw or None,
                         amount_min=float(amt_min) if amt_min else None,
                         amount_max=float(amt_max) if amt_max else None,
                         date_from=df or None, date_to=dt or None,
                         category=cat or None, page=page, per_page=per)
    # export
    if not res["items"]:
        console.print("[yellow]No results[/yellow]")
        return
    # ask export options
    ex = Prompt.ask("Export as (none/excel/pdf/both)", choices=["none","excel","pdf","both"], default="none")
    if ex in ("excel","both"):
        export_to_excel_rows(res["items"])
    if ex in ("pdf","both"):
        embed = Confirm.ask("Embed trend chart in PDF?", default=True)
        export_to_pdf_rows(res["items"], embed_chart=embed)

def monthly_category_summary():
    """Print monthly total per category for current month."""
    session = SessionLocal()
    try:
        today = datetime.date.today()
        month_start = today.replace(day=1)
        import calendar
        last_day = calendar.monthrange(today.year, today.month)[1]
        month_end = today.replace(day=last_day)

        # query totals per category
        results = session.query(
            Category.name,
            func.sum(Expense.amount)
        ).join(Expense, Expense.category_id==Category.id)\
         .filter(Expense.deleted==False,
                 Expense.date >= month_start,
                 Expense.date <= month_end)\
         .group_by(Category.id)\
         .order_by(Category.name).all()

        table = Table(title=f"Category Totals - {today.strftime('%B %Y')}", box=box.SIMPLE, show_header=True, header_style="bold magenta")
        table.add_column("Category")
        table.add_column("Total (₹)", justify="right")
        total_month = 0.0
        for cat, amt in results:
            total_month += amt or 0.0
            table.add_row(cat, f"{amt or 0.0:.2f}")
        console.print(table)
        console.print(f"[cyan]Total all categories: ₹{total_month:.2f}[/cyan]")
    finally:
        session.close()


def prompt_budget():
    # create or update budget for category
    cat = Prompt.ask("Category for budget").strip().title()
    try:
        amt = float(Prompt.ask("Monthly budget amount (₹)"))
    except:
        console.print("[red]Invalid amount[/red]")
        return
    curr = Prompt.ask("Currency", default="INR")
    session = SessionLocal()
    try:
        c = create_category_if_missing(session, cat)
        # upsert budget (one budget per category used here)
        b = session.query(Budget).filter(Budget.category_id == c.id).first()
        if b:
            b.amount = amt
            b.currency = curr
        else:
            b = Budget(category_id=c.id, amount=amt, currency=curr)
            session.add(b)
        session.commit()
        console.print(f"[green]Budget set for {c.name}: ₹{amt:.2f} {curr}[/green]")
    except Exception as exc:
        session.rollback()
        console.print(f"[red]Error: {exc}[/red]")
    finally:
        session.close()

def prompt_trend():
    days = int(Prompt.ask("Days for trend", default="30"))
    data = spending_trend(days)
    # simple ASCII table
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("Date")
    table.add_column("Total (₹)", justify="right")
    for d,t in data:
        table.add_row(d, f"{t:.2f}")
    console.print(table)
    # show small line chart using matplotlib inline popup saved to temp file
    if Confirm.ask("Show chart image and export to PDF?", default=True):
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        dates = [d for d,_ in data]
        totals = [t for _,t in data]
        plt.figure(figsize=(8,2.5))
        plt.plot(dates, totals, marker='o', linewidth=1)
        plt.xticks(rotation=45, fontsize=6)
        plt.tight_layout()
        plt.savefig(tmp.name, dpi=150)
        plt.close()
        console.print(f"[green]Chart saved to {tmp.name}[/green]")
        if Confirm.ask("Attach chart to new PDF report?", default=True):
            # generate PDF with chart only
            export_to_pdf_rows([], filename=f"trend_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", embed_chart=True)

# ---------- Main Menu ----------
def main_menu():
    while True:
        console.print(Panel("Expense Manager PRO (SQLite + SQLAlchemy)", style="bold white on blue"))
        console.print("1. Add Expense")
        console.print("2. View Expenses (pagination)")
        console.print("3. Soft Delete Expense")
        console.print("4. Update Expense")
        console.print("5. Monthly Total")
        console.print("6. Monthly Category Breakdown")
        console.print("7. Filter by Category")
        console.print("8. Filter by Date")
        console.print("9. Search & Export")
        console.print("10. Set Category Budget")
        console.print("11. Spending Trend (days)")
        console.print("12. Export All to Excel")
        console.print("13. Export All to PDF (with chart)")
        console.print("14. Exit")

        choice = Prompt.ask("Choose", choices=[str(i) for i in range(1,15)])
        if choice == "1":
            prompt_add()
        elif choice == "2":
            prompt_view()
        elif choice == "3":
            prompt_delete()
        elif choice == "4":
            prompt_update()
        elif choice == "5":
            # monthly total
            session = SessionLocal()
            try:
                now = datetime.date.today()
                s = session.query(func.sum(Expense.amount)).filter(
                    Expense.deleted==False,
                    Expense.date >= now.replace(day=1),
                    Expense.date <= now
                ).scalar() or 0.0
                console.print(Panel(f"Total this month: ₹{float(s):.2f}", style="cyan"))
            finally:
                session.close()
        elif choice == "6":
            monthly_category_summary()
        elif choice == "7":
            cat = Prompt.ask("Category").strip()
            res = search_expenses(category=cat, per_page=1000)
            if res["items"]:
                export = Confirm.ask("Export results?", default=False)
                show_table = Table(box=box.SIMPLE)
                show_table.add_column("ID"); show_table.add_column("Amount"); show_table.add_column("Category"); show_table.add_column("Date"); show_table.add_column("Note")
                for r in res["items"]:
                    show_table.add_row(str(r["id"]), f"{r['amount']:.2f}", r['category'], r['date'], str(r['note'])[:30])
                console.print(show_table)
                if export:
                    ex = Prompt.ask("Export as excel/pdf/both/none", choices=["excel","pdf","both","none"], default="none")
                    if ex in ("excel","both"):
                        export_to_excel_rows(res["items"])
                    if ex in ("pdf","both"):
                        export_to_pdf_rows(res["items"], embed_chart=True)
            else:
                console.print("[yellow]No items[/yellow]")
        elif choice == "8":
            date = Prompt.ask("Date (YYYY-MM-DD)")
            d = valid_date_str(date)
            if not d:
                console.print("[red]Invalid date[/red]")
            else:
                res = search_expenses(date_from=d, date_to=d, per_page=1000)
                if res["items"]:
                    export_to_excel_rows(res["items"])
                else:
                    console.print("[yellow]No items on that date[/yellow]")
        elif choice == "9":
            prompt_search_export()
        elif choice == "10":
            prompt_budget()
        elif choice == "11":
            prompt_trend()
        elif choice == "12":
            # export all
            res = search_expenses(per_page=10000)
            export_to_excel_rows(res["items"])
        elif choice == "13":
            res = search_expenses(per_page=10000)
            export_to_pdf_rows(res["items"], embed_chart=True)
        elif choice == "14":
            console.print("[green]Goodbye[/green]")
            break

if __name__ == "__main__":
    try:
        init_db()
        main_menu()
    except KeyboardInterrupt:
        console.print("\n[red]Interrupted. Exiting.[/red]")
        sys.exit(0)
 